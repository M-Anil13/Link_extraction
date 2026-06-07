import argparse
import queue
import re
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

# --- Selectors -------------------------------------------------------------
# We match by accessible button NAME with anchored regex instead of substring
# text. Substring 'Apply' also matched 'Applied', so an already-applied card
# looked like a fresh job and fed the stuck-card loop. Anchored regex avoids it.
# If you inspect Jobright's DOM and find stable data-testid/class hooks, swap
# these helpers to page.locator("[data-testid='...']") for resilience.
APPLY_RE = re.compile(r"^\s*Apply( Now| With Autofill)?\s*$", re.I)
MANUAL_RE = re.compile(r"^\s*(No, )?Apply Manually|Apply Without Customizing\s*$", re.I)
YES_APPLIED_RE = re.compile(r"^\s*Yes, I applied!?|I applied\s*$", re.I)
# Logged-out signal: auth CTAs only appear before login (substring match).
LOGGED_OUT_RE = re.compile(
    r"sign\s*in|sign\s*up|log\s*in|login|get started|continue with", re.I
)


def apply_buttons(page):
    """All fresh Apply buttons in the feed (excludes 'Applied')."""
    return page.get_by_role("button", name=APPLY_RE)


def manual_buttons(page):
    return page.get_by_role("button", name=MANUAL_RE)


def yes_applied_buttons(page):
    return page.get_by_role("button", name=YES_APPLIED_RE)

# Edit these defaults if you want to control behavior directly in code.
DEFAULT_PROFILE = "vamshi"
DEFAULT_MAX_LINKS = 31
DEFAULT_START_URL = "https://jobright.ai"
DEFAULT_OUTPUT_FILE = "filtered_job_links.xlsx"

# Short, explicit timeouts (ms). Default Playwright load wait is 30s -> hangs.
POPUP_TIMEOUT = 6000        # wait for a new tab to open after manual apply
LOAD_TIMEOUT = 8000         # wait for a page to finish loading
MANUAL_BTN_TIMEOUT = 600    # wait for the "Apply Manually" button
MODAL_SETTLE = 600          # small settle after opening a modal

# Interactive login (CDP screencast) — fixed viewport so frame pixels == page
# coordinates 1:1, making click/key replay from the browser accurate.
LOGIN_VIEWPORT = {"width": 1280, "height": 800}
LOGIN_TIMEOUT_S = 300       # max seconds to wait for the user to finish login

# A single feed card can stall (embed with no new tab, confirm button missing).
# After this many no-progress passes on the same first card, force past it.
MAX_STUCK_PER_CARD = 3

BLOCK_KEYWORDS = [
    "Security Clearance",
    "U.S. Citizen Only",
]

BLOCK_PORTAL_DOMAINS = [
    "linkedin.com",
    "glassdoor.com",
    "monster.com",
    "ziprecruiter.com",
    "jobright.ai",
    "simplyhired.com",
    "careerbuilder.com",
    "hackajob.com",
]

ATS_DOMAINS = [
    "myworkdayjobs.com",
    "workday.com",
    "greenhouse.io",
    "boards.greenhouse.io",
    "lever.co",
    "jobs.lever.co",
    "icims.com",
    "smartrecruiters.com",
    "successfactors.com",
    "taleo.net",
    "myworkday.com",
    "jobvite.com",
    "bamboohr.com",
    "recruitee.com",
    "applicantpro.com",
    "brassring.com",
    "paylocity.com",
    "workforcenow.adp.com",
    "oraclecloud.com",
    "dayforcehcm.com",
    "ceridian.com",
    "ats.rippling.com",
    "rippling.com",
]


def is_application_url(url):
    if not url:
        return False
    u = url.lower()
    if any(domain in u for domain in BLOCK_PORTAL_DOMAINS):
        return False
    if any(domain in u for domain in ATS_DOMAINS):
        return True
    if "/apply" in u or "application" in u:
        return True

    # Practical fallback: accept direct career/job posting paths.
    job_path_markers = [
        "/jobs/",
        "/job/",
        "/careers/",
        "/positions/",
        "/openings/",
        "/p/",
    ]
    if any(marker in u for marker in job_path_markers):
        return True
    return False


def safe_wait_load(pg, timeout=LOAD_TIMEOUT):
    """wait_for_load_state with a bounded timeout. Never raises.

    The original code called wait_for_load_state() with the default 30s and no
    try/except, so a blank/slow popup hung the whole run and then crashed.
    """
    try:
        pg.wait_for_load_state("domcontentloaded", timeout=timeout)
    except Exception:
        pass


def load_existing(output_file):
    """Resume: read previously saved URLs from a sheet (per-session, no globals).

    Returns (collected_links list, saved_links set). Safe if file missing.
    """
    collected, saved = [], set()
    path = Path(output_file)
    if not path.exists():
        return collected, saved
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            url = row[1]
            if url and url not in saved:
                saved.add(url)
                collected.append(url)
        wb.close()
        print(f"Resume: loaded {len(collected)} existing links from {output_file}")
    except Exception as e:
        print("Resume: could not read existing file -", repr(e))
    return collected, saved


def save_to_excel(output_file, collected_links):
    path = Path(output_file)
    if path.parent and not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Job Links"
    ws.append(["S.No", "Job URL"])

    for i, link in enumerate(collected_links, start=1):
        ws.append([i, link])

    wb.save(output_file)
    print(f"Saved {output_file}")


def try_load_more_jobs(page, previous_total):
    """Trigger infinite scroll and return True when more jobs are loaded."""
    try:
        apply_buttons(page).last.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    page.mouse.wheel(0, 2600)
    page.wait_for_timeout(1200)

    new_total = apply_buttons(page).count()
    print(f"DEBUG: jobs before scroll={previous_total}, after scroll={new_total}")
    return new_total > previous_total


def click_yes_applied_fast(page):
    """Click Jobright's 'Yes, I applied!' confirm so the card leaves the feed."""
    try:
        btn = yes_applied_buttons(page).first
        if btn.count() > 0:
            btn.click(force=True, timeout=800)
            return True
    except Exception:
        pass
    return False


def dismiss_modal(page):
    """Best-effort close of any open Jobright modal."""
    if not click_yes_applied_fast(page):
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
    page.wait_for_timeout(400)


def extract_external_url(context, page):
    """Return the external application URL after clicking manual apply.

    Handles three Jobright flows, all bounded by short timeouts:
      1. New tab opens (workday / icims / rippling)         -> read new tab url
      2. Inline embed, no new tab (greenhouse /embed/...)   -> read current url
      3. Same-tab navigation away from jobright             -> read page url
    """
    manual_apply_button = manual_buttons(page)

    # Flow 1: manual button present -> expect a popup tab.
    try:
        manual_apply_button.first.wait_for(timeout=MANUAL_BTN_TIMEOUT)
        try:
            with context.expect_page(timeout=POPUP_TIMEOUT) as new_page_info:
                manual_apply_button.first.click(force=True)
            new_page = new_page_info.value
            safe_wait_load(new_page)
            return new_page.url
        except Exception:
            # Button clicked but no popup (inline embed / same-tab nav).
            page.wait_for_timeout(MODAL_SETTLE)
    except Exception:
        # No manual button at all.
        pass

    # Flow 2/3: check for a stray extra tab, else current page if it left jobright.
    if len(context.pages) > 1:
        new_page = context.pages[-1]
        safe_wait_load(new_page)
        url = new_page.url
        if "jobright.ai" not in (url or "").lower():
            return url
    if "jobright.ai" not in (page.url or "").lower():
        return page.url
    return None


def close_extra_tabs(context, keep_page):
    """Close leftover application tabs so they don't pile up / get re-read."""
    for pg in list(context.pages):
        if pg is keep_page:
            continue
        try:
            pg.close()
        except Exception:
            pass


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract external job application links from Jobright.ai."
    )
    parser.add_argument(
        "--profile",
        default=DEFAULT_PROFILE,
        help=(
            "Profile name or path for persistent Chrome data. "
            "If name is provided, folder is created inside ./chrome-profiles/."
        ),
    )
    parser.add_argument(
        "--max-links",
        type=int,
        default=DEFAULT_MAX_LINKS,
        help="Maximum number of unique application links to save.",
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_START_URL,
        help="Starting URL (default: https://jobright.ai).",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_FILE,
        help="Excel output file name/path.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run browser in headless mode.",
    )
    return parser.parse_args()


def resolve_profile_dir(profile_arg):
    profile_path = Path(profile_arg)
    if profile_path.is_absolute() or profile_path.parent != Path("."):
        target = profile_path
    else:
        target = Path("chrome-profiles") / profile_arg

    target.mkdir(parents=True, exist_ok=True)
    return str(target.resolve())


def logged_out_markers(page):
    """True if auth CTAs (Sign in / Log in / Get started) are on the page."""
    try:
        if page.get_by_role("button", name=LOGGED_OUT_RE).count() > 0:
            return True
        if page.get_by_role("link", name=LOGGED_OUT_RE).count() > 0:
            return True
    except Exception:
        pass
    return False


def is_logged_in(page, wait_ms=5000):
    """Logged in = jobs feed Apply buttons present AND no auth CTA visible.

    The public landing page has a marketing 'Apply Now' button, so checking
    Apply buttons alone gives a false positive; the auth-CTA check rules it out.
    """
    if logged_out_markers(page):
        return False
    try:
        apply_buttons(page).first.wait_for(timeout=wait_ms)
        return not logged_out_markers(page)
    except Exception:
        return False


def interactive_login(context, page, emit, input_queue, stop_requested,
                      timeout_s=LOGIN_TIMEOUT_S):
    """Stream the page to the UI (CDP screencast) and replay user input until
    login completes. Returns True once the jobs feed appears or the user
    signals done. No VNC/Docker needed — pure CDP.
    """
    cdp = context.new_cdp_session(page)
    frames = queue.Queue()

    def on_frame(params):
        frames.put(params)

    cdp.on("Page.screencastFrame", on_frame)
    cdp.send("Page.startScreencast", {
        "format": "jpeg", "quality": 55,
        "maxWidth": LOGIN_VIEWPORT["width"], "maxHeight": LOGIN_VIEWPORT["height"],
        "everyNthFrame": 1,
    })
    emit("need_login", {"message": "Log in to Jobright in the panel.",
                        "width": LOGIN_VIEWPORT["width"],
                        "height": LOGIN_VIEWPORT["height"]})

    start = time.time()
    done = False
    try:
        while time.time() - start < timeout_s and not stop_requested():
            page.wait_for_timeout(120)  # pump CDP events + input

            # Push the most recent frame to the UI (drop stale ones).
            last = None
            while not frames.empty():
                params = frames.get()
                last = params.get("data")
                try:
                    cdp.send("Page.screencastFrameAck",
                             {"sessionId": params.get("sessionId")})
                except Exception:
                    pass
            if last:
                emit("frame", {"data": last})

            # Replay queued user input from the browser.
            while input_queue is not None and not input_queue.empty():
                ev = input_queue.get()
                etype = ev.get("type")
                try:
                    if etype == "click":
                        page.mouse.click(ev["x"], ev["y"])
                    elif etype == "move":
                        page.mouse.move(ev["x"], ev["y"])
                    elif etype == "scroll":
                        page.mouse.wheel(0, ev.get("dy", 0))
                    elif etype == "char":
                        page.keyboard.type(ev["value"])
                    elif etype == "key":
                        page.keyboard.press(ev["value"])
                    elif etype == "login_done":
                        done = True
                except Exception as e:
                    emit("error", {"message": f"input replay: {e!r}"})
                if done:
                    break

            # Quick, non-blocking completion check (avoid 5s wait per tick).
            if done or (not logged_out_markers(page)
                        and apply_buttons(page).count() > 0):
                done = True
                break
    finally:
        try:
            cdp.send("Page.stopScreencast")
        except Exception:
            pass

    if done:
        emit("login_ok", {"message": "Login detected. Starting extraction..."})
    return done


def extract_links(
    profile=DEFAULT_PROFILE,
    max_links=DEFAULT_MAX_LINKS,
    url=DEFAULT_START_URL,
    output=DEFAULT_OUTPUT_FILE,
    headless=False,
    login_wait=5000,
    interactive=False,
    input_queue=None,
    on_event=None,
    should_stop=None,
):
    """Core extraction. Emits structured events via on_event(type, payload).

    Event types: status, found, link, skip, done, error, stopped.
    `should_stop()` (optional) is polled each loop so a UI can cancel.
    Runs synchronously; call from a worker thread when used inside async code.
    """
    def emit(etype, payload=None):
        print(f"[{etype}] {payload if payload is not None else ''}")
        if on_event:
            try:
                on_event(etype, payload or {})
            except Exception:
                pass

    def stop_requested():
        return bool(should_stop and should_stop())

    # Per-run state (no module globals -> safe for concurrent sessions).
    collected_links, saved_links = load_existing(output)
    emit("status", {"message": "Resumed existing links",
                    "links": list(collected_links)})

    profile_dir = resolve_profile_dir(profile)

    # Interactive mode streams the page to the UI, so it must run headless with
    # a fixed viewport (frame pixels map 1:1 to page coordinates).
    if interactive:
        launch_kwargs = dict(headless=True, viewport=LOGIN_VIEWPORT)
    else:
        launch_kwargs = dict(headless=headless, args=["--start-maximized"])

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=profile_dir,
            channel="chrome",
            **launch_kwargs,
        )

        page = context.pages[0] if context.pages else context.new_page()
        page.goto(url)
        emit("status", {"message": f"Opened {url} (profile: {profile})"})

        if not is_logged_in(page):
            if interactive:
                ok = interactive_login(context, page, emit, input_queue,
                                       stop_requested)
                if not ok:
                    emit("error", {"message": "Login not completed."})
                    context.close()
                    emit("done", {"saved": 0, "attempts": 0,
                                  "total": len(collected_links)})
                    return list(collected_links)
            else:
                # Non-interactive: give a visible browser time for manual login.
                page.wait_for_timeout(login_wait)
                if not is_logged_in(page):
                    emit("error", {"message": "Not logged in (no jobs found)."})
                    context.close()
                    emit("done", {"saved": 0, "attempts": 0,
                                  "total": len(collected_links)})
                    return list(collected_links)

        emit("status", {"message": "Jobs detected. Extracting..."})

        saved_count = 0
        attempts = 0
        no_new_job_rounds = 0
        max_no_new_job_rounds = 5
        stuck_on_card = 0

        while saved_count < max_links:
            if stop_requested():
                emit("stopped", {"saved": saved_count})
                break

            total = apply_buttons(page).count()

            if total == 0:
                loaded = try_load_more_jobs(page, total)
                if loaded:
                    no_new_job_rounds = 0
                    continue
                no_new_job_rounds += 1
                if no_new_job_rounds >= max_no_new_job_rounds:
                    emit("status", {"message": "No more jobs after scrolling."})
                    break
                continue

            no_new_job_rounds = 0
            attempts += 1
            emit("progress", {"attempt": attempts, "saved": saved_count,
                              "max": max_links, "stuck": stuck_on_card})

            made_progress = False
            try:
                apply_btn = apply_buttons(page).first
                apply_btn.click()
                page.wait_for_timeout(MODAL_SETTLE)

                page_text = page.inner_text("body")
                if any(keyword in page_text for keyword in BLOCK_KEYWORDS):
                    emit("skip", {"reason": "restricted"})
                    dismiss_modal(page)
                    made_progress = True
                else:
                    apply_buttons(page).first.click(force=True)
                    page.wait_for_timeout(MODAL_SETTLE)

                    job_url = extract_external_url(context, page)
                    page.bring_to_front()

                    if job_url:
                        if not is_application_url(job_url):
                            emit("skip", {"reason": "portal", "url": job_url})
                            dismiss_modal(page)
                            made_progress = True
                        elif job_url in saved_links:
                            emit("skip", {"reason": "duplicate", "url": job_url})
                            dismiss_modal(page)
                            made_progress = True
                        else:
                            saved_links.add(job_url)
                            collected_links.append(job_url)
                            save_to_excel(output, collected_links)
                            saved_count += 1
                            emit("link", {"url": job_url, "index": saved_count,
                                          "max": max_links})
                            dismiss_modal(page)
                            made_progress = True
                    else:
                        emit("skip", {"reason": "no-url"})
                        dismiss_modal(page)
            except Exception as e:
                emit("error", {"message": repr(e)})
                dismiss_modal(page)

            close_extra_tabs(context, page)
            page.wait_for_timeout(400)

            if made_progress:
                stuck_on_card = 0
            else:
                stuck_on_card += 1
                if stuck_on_card >= MAX_STUCK_PER_CARD:
                    emit("status", {"message": "Card stuck - scrolling past."})
                    page.mouse.wheel(0, 1200)
                    page.wait_for_timeout(800)
                    stuck_on_card = 0

        context.close()

    save_to_excel(output, collected_links)
    emit("done", {"saved": saved_count, "attempts": attempts,
                  "total": len(collected_links)})
    return list(collected_links)


def run():
    args = parse_args()
    extract_links(
        profile=args.profile,
        max_links=args.max_links,
        url=args.url,
        output=args.output,
        headless=args.headless,
    )


if __name__ == "__main__":
    run()
